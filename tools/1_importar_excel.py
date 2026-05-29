#!/usr/bin/env python3
"""
PASSO 1 — Importar tabela de preços de um arquivo Excel para o banco local.

Uso:
    python3 1_importar_excel.py

O script vai perguntar interativamente:
  - Caminho do arquivo Excel
  - Aba com os produtos (ex: Feminina, Masculina)
  - Nome da fábrica
  - Nome da tabela de preço
  - Coleção, estação, ano
  - Regras de comissão (ou reutiliza de tabela existente)

Formato esperado do Excel:
  Linha 4: cabeçalho  → col0=categoria, col4="VALOR", col5-7=descontos negativos, col8="OBSERVAÇÃO"
  Linha 5+: produtos  → col0=ref, col1=nome, col2=modelo, col3=grade, col4=preço base, col8=observação

  Aba "Packs" (opcional): grades detalhadas dos packs PKTE
    - Linha com "Ref. PKTE####" → inicia um pack
    - Próxima linha: "cor", tamanho1, tamanho2, ...
    - Linhas seguintes: cor, qtd1, qtd2, ...
"""

import sys
import re
import uuid
import subprocess
import openpyxl
from config import PSQL_PATH, DB_LOCAL, DB_LOCAL_USER

# ── helpers de banco ──────────────────────────────────────────────────────────

def db_run(sql: str, args: list = None):
    """Executa SQL no banco local (sem retorno)."""
    cmd = [PSQL_PATH, "-U", DB_LOCAL_USER, "-d", DB_LOCAL, "-c", sql]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"Erro SQL:\n{r.stderr}\n\nSQL:\n{sql[:300]}")

def db_query(sql: str) -> str:
    """Executa SQL e retorna saída como string."""
    cmd = [PSQL_PATH, "-U", DB_LOCAL_USER, "-d", DB_LOCAL, "-t", "-A", "-c", sql]
    r = subprocess.run(cmd, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"Erro SQL:\n{r.stderr}")
    return r.stdout.strip()

def db_rows(sql: str) -> list[list[str]]:
    """Retorna lista de listas com o resultado."""
    out = db_query(sql)
    if not out:
        return []
    return [row.split('|') for row in out.split('\n') if row]

# ── leitura do Excel ──────────────────────────────────────────────────────────

def parse_price_sheet(ws) -> list[dict]:
    """
    Lê aba de preços (Feminina/Masculina/etc) e retorna lista de produtos.
    Ignora linhas sem referência válida (TE#### ou PKTE####).
    """
    products = []
    for row in ws.iter_rows(values_only=True):
        ref = str(row[0] or '').strip().upper()
        if not re.match(r'^(TE|PKTE)\d+$', ref):
            continue
        name    = str(row[1] or '').strip()
        model   = str(row[2] or '').strip()
        grade   = str(row[3] or '').strip()
        price   = row[4]
        obs     = str(row[8] or '').strip() if len(row) > 8 else ''

        if price is None or not isinstance(price, (int, float)):
            continue
        price = abs(float(price))
        if price <= 0:
            continue

        ptype = 'pack' if ref.startswith('PKTE') else 'regular'
        products.append({
            'reference':    ref,
            'product_name': name,
            'model':        model,
            'size_range':   grade,
            'base_price':   round(price, 2),
            'type':         ptype,
            'observation':  obs,
        })
    return products


def parse_packs_sheet(ws) -> dict[str, list[dict]]:
    """
    Lê aba 'Packs' e retorna dict: { 'PKTE11375': [grade_configs...], ... }

    Formato:
      Linha com "Ref. PKTE####" → início de um pack
      Próxima linha não-vazia: "cor", tam1, tam2, ...  → cabeçalho de tamanhos
      Linhas seguintes: cor, qtd1, qtd2, ...            → cores
      Linha em branco ou próximo Ref. → fim do pack
    """
    result: dict[str, list] = {}
    current_ref = None
    size_headers = []
    grades = []
    in_header_next = False

    for row in ws.iter_rows(values_only=True):
        row = list(row)
        row_str = ' '.join(str(c) for c in row if c is not None)

        # Detecta nova referência
        ref_match = re.search(r'(PKTE\d+)', row_str, re.IGNORECASE)
        if ref_match:
            if current_ref and grades:
                result[current_ref] = grades
            current_ref = ref_match.group(1).upper()
            size_headers = []
            grades = []
            in_header_next = True
            continue

        if current_ref is None:
            continue

        # Linha de cabeçalho de tamanhos (primeira linha após o Ref.)
        if in_header_next:
            non_none = [c for c in row if c is not None]
            if not non_none:
                continue
            first = str(non_none[0]).strip().lower()
            if first in ('cor', 'color', 'cores'):
                size_headers = [str(c) for c in row[1:] if c is not None and str(c).strip() not in ('', 'None')]
                in_header_next = False
                continue

        # Linha de cor + quantidades
        if not in_header_next and size_headers:
            cor = str(row[0] or '').strip()
            if not cor or cor.lower() in ('none', ''):
                continue
            # Pega as quantidades nas posições dos cabeçalhos
            qtds_raw = row[1:len(size_headers)+1]
            sizes = {}
            total = 0
            for h, q in zip(size_headers, qtds_raw):
                v = int(q) if q and str(q).strip() not in ('', 'None') else 0
                sizes[h] = v
                total += v
            if total > 0:
                grades.append({
                    'color':        cor,
                    'sizes':        sizes,
                    'total_pieces': total,
                    'sort_order':   len(grades),
                })

    # Última referência
    if current_ref and grades:
        result[current_ref] = grades

    return result

# ── funções de banco ──────────────────────────────────────────────────────────

def get_or_create_factory(name: str) -> str:
    rows = db_rows(f"SELECT id FROM factories WHERE name ILIKE '{name}'")
    if rows:
        return rows[0][0]
    new_id = str(uuid.uuid4())
    db_run(f"INSERT INTO factories (id, name) VALUES ('{new_id}', '{name.replace(chr(39), chr(39)*2)}')")
    print(f"  → Fábrica criada: {name}")
    return new_id

def create_price_table(factory_id: str, name: str, collection: str, season: str, year: int) -> str:
    new_id = str(uuid.uuid4())
    db_run(f"""
        INSERT INTO price_tables (id, factory_id, name, collection, season, year, active)
        VALUES ('{new_id}', '{factory_id}',
                '{name.replace(chr(39), chr(39)*2)}',
                '{collection.replace(chr(39), chr(39)*2)}',
                '{season.replace(chr(39), chr(39)*2)}',
                {year}, true)
    """)
    print(f"  → Tabela criada: {name}")
    return new_id

def create_commission_rules(price_table_id: str, rules: list[tuple]):
    """
    rules: [(discount_pct, total_pct, rep_pct, office_pct), ...]
    """
    for i, (disc, total, rep, office) in enumerate(rules):
        new_id = str(uuid.uuid4())
        db_run(f"""
            INSERT INTO discount_commission_rules
              (id, price_table_id, discount_pct, total_commission_pct, rep_commission_pct, office_commission_pct, sort_order)
            VALUES ('{new_id}', '{price_table_id}', {disc}, {total}, {rep}, {office}, {i})
        """)
    print(f"  → {len(rules)} regras de comissão criadas")

def insert_products(price_table_id: str, products: list[dict], pack_grades: dict[str, list]) -> tuple[int, int]:
    inserted = skipped = 0
    for p in products:
        ref = p['reference']
        # Verifica se já existe
        exists = db_query(f"SELECT id FROM products WHERE price_table_id='{price_table_id}' AND reference='{ref}'")
        if exists:
            skipped += 1
            continue

        prod_id = str(uuid.uuid4())
        name_esc  = p['product_name'].replace("'", "''")
        model_esc = p['model'].replace("'", "''")
        obs_esc   = p['observation'].replace("'", "''")
        db_run(f"""
            INSERT INTO products
              (id, price_table_id, reference, product_name, model, size_range, base_price, type, observation, active)
            VALUES (
              '{prod_id}', '{price_table_id}', '{ref}',
              '{name_esc}', '{model_esc}',
              '{p['size_range']}', {p['base_price']},
              '{p['type']}', '{obs_esc}', true
            )
        """)

        # Grade de packs
        if p['type'] == 'pack' and ref in pack_grades:
            for gc in pack_grades[ref]:
                gc_id = str(uuid.uuid4())
                # Serializa sizes como JSON
                sizes_json = '{' + ','.join(f'"{k}":{v}' for k, v in gc['sizes'].items()) + '}'
                color_esc = gc['color'].replace("'", "''")
                db_run(f"""
                    INSERT INTO grade_configs (id, product_id, color, sizes, total_pieces, sort_order)
                    VALUES ('{gc_id}', '{prod_id}', '{color_esc}', '{sizes_json}', {gc['total_pieces']}, {gc['sort_order']})
                """)

        inserted += 1

    return inserted, skipped

# ── entrada interativa ────────────────────────────────────────────────────────

def ask(prompt: str, default: str = '') -> str:
    suffix = f' [{default}]' if default else ''
    val = input(f'{prompt}{suffix}: ').strip()
    return val if val else default

def ask_float(prompt: str, default: float) -> float:
    while True:
        val = ask(prompt, str(default))
        try:
            return float(val.replace(',', '.'))
        except ValueError:
            print("  → Valor inválido. Digite um número (ex: 4.11)")

def main():
    print("=" * 60)
    print("  SOMMA PEDIDOS — Importar tabela de preços (Excel)")
    print("=" * 60)
    print()

    # 1. Arquivo Excel
    excel_path = ask("Caminho do arquivo Excel")
    if not excel_path:
        print("Caminho obrigatório.")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Erro ao abrir Excel: {e}")
        sys.exit(1)

    print(f"\nAbas disponíveis: {wb.sheetnames}")
    sheet_name = ask("Nome da aba com os produtos", wb.sheetnames[0])
    if sheet_name not in wb.sheetnames:
        print(f"Aba '{sheet_name}' não encontrada.")
        sys.exit(1)

    has_packs_sheet = 'Packs' in wb.sheetnames or 'PACKS' in wb.sheetnames
    packs_sheet_name = 'Packs' if 'Packs' in wb.sheetnames else ('PACKS' if 'PACKS' in wb.sheetnames else None)

    # 2. Informações da tabela
    print()
    factory_name = ask("Nome da fábrica", "TEEZZ")
    table_name   = ask("Nome da tabela de preço")
    collection   = ask("Coleção (ex: Inverno 2026 Feminina)")
    season       = ask("Estação (ex: Inverno)", "Inverno")
    year_str     = ask("Ano", "2026")
    year = int(year_str)

    # 3. Regras de comissão
    print()
    print("Regras de comissão (desconto → % total | % vendedor | % escritório)")
    print("Pressione Enter para usar os padrões TEEZZ:")
    print("  0.00% → 10.00 | 6.00 | 4.00")
    print("  4.11% →  8.00 | 4.50 | 3.50")
    print("  6.85% →  7.00 | 4.00 | 3.00")
    print(" 10.00% →  6.00 | 3.00 | 3.00")

    use_default = ask("Usar regras padrão TEEZZ? (s/n)", "s").lower().startswith('s')

    if use_default:
        rules = [
            (0.00, 10.00, 6.00, 4.00),
            (4.11,  8.00, 4.50, 3.50),
            (6.85,  7.00, 4.00, 3.00),
            (10.00, 6.00, 3.00, 3.00),
        ]
    else:
        rules = []
        print("Digite as regras (deixe vazio para terminar):")
        while True:
            disc  = ask("  Desconto %", "")
            if not disc:
                break
            disc  = float(disc.replace(',', '.'))
            total = ask_float("    Total comissão %", 10.00)
            rep   = ask_float("    Vendedor %", 6.00)
            office= ask_float("    Escritório %", 4.00)
            rules.append((disc, total, rep, office))

    # ── Processamento ────────────────────────────────────────────────────────

    print()
    print("Lendo Excel...")
    ws = wb[sheet_name]
    products = parse_price_sheet(ws)
    print(f"  → {len(products)} produtos encontrados na aba '{sheet_name}'")

    pack_grades = {}
    if has_packs_sheet:
        ws_packs = wb[packs_sheet_name]
        pack_grades = parse_packs_sheet(ws_packs)
        print(f"  → {len(pack_grades)} packs com grade lidos da aba '{packs_sheet_name}'")

    packs_in_sheet = [p for p in products if p['type'] == 'pack']
    regs_in_sheet  = [p for p in products if p['type'] == 'regular']
    packs_with_grade = [p for p in packs_in_sheet if p['reference'] in pack_grades]
    print(f"     {len(regs_in_sheet)} regulares + {len(packs_in_sheet)} packs ({len(packs_with_grade)} com grade)")

    # Confirmação
    print()
    print("─" * 60)
    print(f"  Fábrica:  {factory_name}")
    print(f"  Tabela:   {table_name}")
    print(f"  Coleção:  {collection} — {season} {year}")
    print(f"  Produtos: {len(products)}")
    print(f"  Regras:   {len(rules)} faixas de desconto")
    print("─" * 60)
    confirm = ask("Confirmar importação? (s/n)", "s").lower()
    if not confirm.startswith('s'):
        print("Cancelado.")
        sys.exit(0)

    print()
    print("Importando...")

    factory_id = get_or_create_factory(factory_name)
    table_id   = create_price_table(factory_id, table_name, collection, season, year)
    create_commission_rules(table_id, rules)
    inserted, skipped = insert_products(table_id, products, pack_grades)

    print()
    print("=" * 60)
    print(f"  ✅ CONCLUÍDO!")
    print(f"  Inseridos: {inserted} | Ignorados (já existiam): {skipped}")
    print(f"  ID da tabela: {table_id}")
    print()
    print("  Próximos passos:")
    print("  1. python3 2_importar_fotos_pdf.py  ← importar fotos dos lookbooks")
    print("  2. python3 3_sincronizar_producao.py ← enviar para o Railway")
    print("=" * 60)

if __name__ == '__main__':
    main()
